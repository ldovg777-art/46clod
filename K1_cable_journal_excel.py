"""
Generate K1 Cable Journal in Excel format
Based on block diagram (Fig. 5) and electrical schemes (421457.103E1)
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ─────────────────────────────────────────────────────────
thin = Side(style='thin')
medium = Side(style='medium')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header = Border(left=thin, right=thin, top=medium, bottom=medium)

title_font = Font(name='Arial', size=14, bold=True)
header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
sub_header_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
small_font = Font(name='Arial', size=9)
note_font = Font(name='Arial', size=10, color='CC0000', bold=True)

header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ekor_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
plant_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
error_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)


def style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border_header


def style_row(ws, row, col_start, col_end, font=normal_font, fill=None,
              align=center):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = align
        cell.border = border_all
        if fill:
            cell.fill = fill


# ═══════════════════════════════════════════════════════════════════
# Sheet 1: КАБЕЛЬНЫЙ ЖУРНАЛ (Cable Schedule)
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Кабельный журнал"
ws1.sheet_properties.tabColor = '2F5496'

# Column widths
col_widths = [6, 28, 28, 24, 10, 10, 24, 16]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title block
ws1.merge_cells('A1:H1')
ws1['A1'] = 'КАБЕЛЬНЫЙ ЖУРНАЛ / CABLE SCHEDULE'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws1.merge_cells('A2:H2')
ws1['A2'] = 'DEEPAK K1 — GP Plant, India  |  Система контроля процессов нейтрализации  |  421457.103E1'
ws1['A2'].font = Font(name='Arial', size=10, italic=True)
ws1['A2'].alignment = Alignment(horizontal='center')

# Headers (row 4)
headers = ['№ каб.', 'Откуда', 'Куда', 'Тип кабеля', 'Сечение', 'Длина, м',
           'Назначение', 'Поставщик']
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 1, 8)

# Cable data
cables = [
    (1, 'Датчик QE1\n(Separator V451)', 'Трансмиттер №1\n(XT1, Sheet 2)',
     '4×1.0 mm² с экраном', '4×1.0', 15,
     'Сигнал датчика\n(RE1, RE2, WE, SE)', 'Ekor'),
    (2, 'Датчик QE2\n(AN tank V452)', 'Трансмиттер №2\n(XT1, Sheet 3)',
     '4×1.0 mm² с экраном', '4×1.0', 15,
     'Сигнал датчика\n(RE1, RE2, WE, SE)', 'Ekor'),
    (3, 'Connection box\n(XT пп.1,3,5,7)', 'Трансмиттер №1\n(XT2, Sheet 2)',
     '2 пары×1.5 mm² вит. пара', '2п×1.5', 15,
     'Питание +24VDC\n+ RS-485', 'K-1 GP Plant'),
    (4, 'Connection box\n(XT пп.2,4,6,8)', 'Трансмиттер №2\n(XT2, Sheet 3)',
     '2 пары×1.5 mm² вит. пара', '2п×1.5', 15,
     'Питание +24VDC\n+ RS-485', 'K-1 GP Plant'),
    (5, 'Трансмиттер №1\n(XT2 пп.5-8, Sheet 2)', 'DCS K-1 GP Plant',
     '2 пары×1.5 mm² с экраном', '2п×1.5', 200,
     '4-20 мА\n(потенциал + pH)', 'K-1 GP Plant'),
    (6, 'Трансмиттер №2\n(XT2 пп.5-8, Sheet 3)', 'DCS K-1 GP Plant',
     '2 пары×1.5 mm² с экраном', '2п×1.5', 200,
     '4-20 мА\n(потенциал + pH)', 'K-1 GP Plant'),
    (7, 'Трансмиттер №1\n(корпус)', 'Шина заземления',
     '1C×2.5 mm² жёлто-зелён.', '1×2.5', 50,
     'Защитное заземление\n(PE)', 'K-1 GP Plant'),
    (8, 'Трансмиттер №2\n(корпус)', 'Шина заземления',
     '1C×2.5 mm² жёлто-зелён.', '1×2.5', 50,
     'Защитное заземление\n(PE)', 'K-1 GP Plant'),
    (9, 'SCADA Ekor Box\n(XT1, Sheet 5)', 'Connection box',
     '2 пары×1.5 mm² вит. пара', '2п×1.5', 250,
     'Питание +24VDC\n+ RS-485 (магистраль)', 'K-1 GP Plant'),
    (10, 'Розетка 230VAC\n(1.5A)', 'SCADA Ekor Box\n(XT2, Sheet 5)',
     '3 жилы×2.5 mm²', '3×2.5', 100,
     'Питание 230VAC\n(L, N, PE)', 'K-1 GP Plant'),
    (11, 'SCADA Ekor Box', 'Operator panel',
     '3×0.75 mm²', '3×0.75', 3,
     'Внутренняя связь', 'Ekor'),
    (12, 'Operator panel', 'LCD Monitor',
     'HDMI', 'HDMI', 3,
     'Видеосигнал', 'Ekor'),
]

for idx, (num, src, dst, ctype, section, length, purpose, supplier) in enumerate(cables):
    row = 5 + idx
    ws1.cell(row=row, column=1, value=num)
    ws1.cell(row=row, column=2, value=src)
    ws1.cell(row=row, column=3, value=dst)
    ws1.cell(row=row, column=4, value=ctype)
    ws1.cell(row=row, column=5, value=section)
    ws1.cell(row=row, column=6, value=length)
    ws1.cell(row=row, column=7, value=purpose)
    ws1.cell(row=row, column=8, value=supplier)

    fill = ekor_fill if supplier == 'Ekor' else plant_fill
    style_row(ws1, row, 1, 8, fill=fill)
    ws1.cell(row=row, column=1).alignment = center
    ws1.cell(row=row, column=5).alignment = center
    ws1.cell(row=row, column=6).alignment = center
    ws1.cell(row=row, column=8).alignment = center
    for c in [2, 3, 4, 7]:
        ws1.cell(row=row, column=c).alignment = left_wrap
    ws1.row_dimensions[row].height = 38

# Totals row
total_row = 5 + len(cables)
ws1.merge_cells(f'A{total_row}:E{total_row}')
ws1.cell(row=total_row, column=1, value='ИТОГО / TOTAL')
ws1.cell(row=total_row, column=1).font = sub_header_font
ws1.cell(row=total_row, column=1).alignment = Alignment(horizontal='right',
                                                         vertical='center')
ws1.cell(row=total_row, column=6, value=f'=SUM(F5:F{total_row-1})')
ws1.cell(row=total_row, column=6).font = sub_header_font
ws1.cell(row=total_row, column=6).alignment = center
style_row(ws1, total_row, 1, 8, fill=sub_header_fill, font=sub_header_font)

# Supplier summary
sr = total_row + 2
ws1.merge_cells(f'A{sr}:C{sr}')
ws1.cell(row=sr, column=1, value='Поставщик / Supplier')
ws1.cell(row=sr, column=4, value='Кабели')
ws1.cell(row=sr, column=5, value='Метраж, м')
style_header_row(ws1, sr, 1, 5)

ws1.merge_cells(f'A{sr+1}:C{sr+1}')
ws1.cell(row=sr+1, column=1, value='Ekor')
ws1.cell(row=sr+1, column=4, value='1, 2, 11, 12')
ws1.cell(row=sr+1, column=5, value=36)
style_row(ws1, sr+1, 1, 5, fill=ekor_fill)

ws1.merge_cells(f'A{sr+2}:C{sr+2}')
ws1.cell(row=sr+2, column=1, value='K-1 GP Plant')
ws1.cell(row=sr+2, column=4, value='3, 4, 5, 6, 7, 8, 9, 10')
ws1.cell(row=sr+2, column=5, value=880)
style_row(ws1, sr+2, 1, 5, fill=plant_fill)

# Legend
lr = sr + 4
ws1.cell(row=lr, column=1, value='Цветовая кодировка:').font = sub_header_font
ws1.cell(row=lr+1, column=1, value='     ')
ws1.cell(row=lr+1, column=1).fill = ekor_fill
ws1.cell(row=lr+1, column=2, value='Поставляет Ekor').font = normal_font
ws1.cell(row=lr+2, column=1, value='     ')
ws1.cell(row=lr+2, column=1).fill = plant_fill
ws1.cell(row=lr+2, column=2, value='Поставляет K-1 GP Plant').font = normal_font


# ═══════════════════════════════════════════════════════════════════
# Sheet 2: СПЕЦИФИКАЦИЯ ЖИЛ (Conductor Assignment)
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Спецификация жил")
ws2.sheet_properties.tabColor = '548235'

col_w2 = [8, 10, 14, 40, 28]
for i, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A1:E1')
ws2['A1'] = 'СПЕЦИФИКАЦИЯ ЖИЛ ПО КАБЕЛЯМ / CONDUCTOR ASSIGNMENT'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

ws2.merge_cells('A2:E2')
ws2['A2'] = 'DEEPAK K1 — GP Plant  |  421457.103E1'
ws2['A2'].font = Font(name='Arial', size=10, italic=True)
ws2['A2'].alignment = Alignment(horizontal='center')

headers2 = ['№ каб.', 'Жила', 'Цвет/Маркир.', 'Назначение', 'Клемма']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, 1, 5)

# Conductor data: (cable_num, conductor, color, purpose, terminal)
conductors = [
    # Cable 1
    (1, '1', 'СИНИЙ (Blue)', 'RE1 — токоотвод электрода сравнения №1', 'XT1 п.1'),
    (1, '2', 'БЕЛЫЙ (White)', 'RE2 — токоотвод электрода сравнения №2', 'XT1 п.2'),
    (1, '3', 'КРАСНЫЙ (Red)', 'WE — токоотвод рабочего электрода', 'XT1 п.3'),
    (1, '4', 'ЧЁРНЫЙ (Black)', 'SE — токоотвод вспомогательного электрода', 'XT1 п.4'),
    (1, 'Экран', '—', 'Shield (заземление)', 'XT1 п.5, п.6'),
    # Cable 2
    (2, '1', 'СИНИЙ (Blue)', 'RE1 — токоотвод электрода сравнения №1', 'XT1 п.1'),
    (2, '2', 'БЕЛЫЙ (White)', 'RE2 — токоотвод электрода сравнения №2', 'XT1 п.2'),
    (2, '3', 'КРАСНЫЙ (Red)', 'WE — токоотвод рабочего электрода', 'XT1 п.3'),
    (2, '4', 'ЧЁРНЫЙ (Black)', 'SE — токоотвод вспомогательного электрода', 'XT1 п.4'),
    (2, 'Экран', '—', 'Shield (заземление)', 'XT1 п.5, п.6'),
    # Cable 3
    (3, 'Пара 1-1', '—', 'U1+ (Supply +24VDC)', 'XT2 п.1 → ConnBox п.1'),
    (3, 'Пара 1-2', '—', 'U1- (Supply -24VDC)', 'XT2 п.2 → ConnBox п.3'),
    (3, 'Пара 2-1', '—', 'Data1+ (RS-485 A)', 'XT2 п.3 → ConnBox п.5'),
    (3, 'Пара 2-2', '—', 'Data1- (RS-485 B)', 'XT2 п.4 → ConnBox п.7'),
    # Cable 4
    (4, 'Пара 1-1', '—', 'U2+ (Supply +24VDC)', 'XT2 п.1 → ConnBox п.2'),
    (4, 'Пара 1-2', '—', 'U2- (Supply -24VDC)', 'XT2 п.2 → ConnBox п.4'),
    (4, 'Пара 2-1', '—', 'Data2+ (RS-485 A)', 'XT2 п.3 → ConnBox п.6'),
    (4, 'Пара 2-2', '—', 'Data2- (RS-485 B)', 'XT2 п.4 → ConnBox п.8'),
    # Cable 5
    (5, 'Пара 1-1', '—', 'T1 Potential+ (4-20 мА, от QS1 OUT)', 'XT2 п.5 → DCS п.1'),
    (5, 'Пара 1-2', '—', 'T1 Potential- (4-20 мА, от QS1 GND)', 'XT2 п.6 → DCS п.2'),
    (5, 'Пара 2-1', '—', 'T1 Conc/pH+ (4-20 мА, от QS2 OUT)', 'XT2 п.7 → DCS п.3'),
    (5, 'Пара 2-2', '—', 'T1 Conc/pH- (4-20 мА, от QS2 GND)', 'XT2 п.8 → DCS п.4'),
    (5, 'Экран', '—', 'Shield', '—'),
    # Cable 6
    (6, 'Пара 1-1', '—', 'T2 Potential+ (4-20 мА, от QS1 OUT)', 'XT2 п.5 → DCS п.5'),
    (6, 'Пара 1-2', '—', 'T2 Potential- (4-20 мА, от QS1 GND)', 'XT2 п.6 → DCS п.6'),
    (6, 'Пара 2-1', '—', 'T2 Conc/pH+ (4-20 мА, от QS2 OUT)', 'XT2 п.7 → DCS п.7'),
    (6, 'Пара 2-2', '—', 'T2 Conc/pH- (4-20 мА, от QS2 GND)', 'XT2 п.8 → DCS п.8'),
    (6, 'Экран', '—', 'Shield', '—'),
    # Cable 7
    (7, '1', 'Жёлто-зелёный', 'PE — защитное заземление Трансмиттера №1', 'Корпус → шина PE'),
    # Cable 8
    (8, '1', 'Жёлто-зелёный', 'PE — защитное заземление Трансмиттера №2', 'Корпус → шина PE'),
    # Cable 9
    (9, 'Пара 1-1', '—', 'U+ (Supply +24VDC)', 'XT1 п.1 → ConnBox пп.1,2'),
    (9, 'Пара 1-2', '—', 'U- (Supply -24VDC)', 'XT1 п.2 → ConnBox пп.3,4'),
    (9, 'Пара 2-1', '—', 'Data+ (RS-485 A)', 'XT1 п.3 → ConnBox пп.5,6'),
    (9, 'Пара 2-2', '—', 'Data- (RS-485 B)', 'XT1 п.4 → ConnBox пп.7,8'),
    # Cable 10
    (10, '1', 'Коричневый', 'L — Фаза 230VAC', 'XT2 L'),
    (10, '2', 'Голубой', 'N — Нейтраль', 'XT2 N'),
    (10, '3', 'Жёлто-зелёный', 'PE — Защитное заземление', 'XT2 PE'),
    # Cable 11
    (11, '1-3', '—', 'Внутренняя связь SCADA → Oper. panel', '—'),
    # Cable 12
    (12, 'HDMI', '—', 'Видеосигнал → LCD Monitor', 'HDMI разъём'),
]

# Color mapping for visual markers
color_map = {
    'СИНИЙ (Blue)': '4472C4',
    'БЕЛЫЙ (White)': 'D9D9D9',
    'КРАСНЫЙ (Red)': 'FF0000',
    'ЧЁРНЫЙ (Black)': '333333',
    'Жёлто-зелёный': 'A9D18E',
    'Коричневый': '8B4513',
    'Голубой': '00B0F0',
}

row = 5
prev_cable = None
for (cab, cond, color, purpose, terminal) in conductors:
    # Add separator between cables
    if cab != prev_cable and prev_cable is not None:
        ws2.cell(row=row, column=1, value='')
        for c in range(1, 6):
            ws2.cell(row=row, column=c).fill = sub_header_fill
            ws2.cell(row=row, column=c).border = border_all
        ws2.row_dimensions[row].height = 6
        row += 1
    prev_cable = cab

    ws2.cell(row=row, column=1, value=cab)
    ws2.cell(row=row, column=2, value=cond)
    ws2.cell(row=row, column=3, value=color)
    ws2.cell(row=row, column=4, value=purpose)
    ws2.cell(row=row, column=5, value=terminal)
    style_row(ws2, row, 1, 5)
    ws2.cell(row=row, column=1).alignment = center
    ws2.cell(row=row, column=2).alignment = center
    ws2.cell(row=row, column=4).alignment = left_wrap
    ws2.cell(row=row, column=5).alignment = left_wrap
    ws2.row_dimensions[row].height = 22

    # Color fill for wire color column
    if color in color_map:
        ws2.cell(row=row, column=3).fill = PatternFill(
            start_color=color_map[color], end_color=color_map[color],
            fill_type='solid')
        if color in ('ЧЁРНЫЙ (Black)',):
            ws2.cell(row=row, column=3).font = Font(name='Arial', size=10,
                                                     color='FFFFFF')

    row += 1


# ═══════════════════════════════════════════════════════════════════
# Sheet 3: КЛЕММНЫЕ КАРТЫ (Terminal Maps)
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Клеммные карты")
ws3.sheet_properties.tabColor = 'BF8F00'

col_w3 = [10, 16, 16, 36]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells('A1:D1')
ws3['A1'] = 'КЛЕММНЫЕ КАРТЫ / TERMINAL MAPS'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

# ── DCS K-1 GP Plant ──
row = 3
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value='DCS K-1 GP Plant (вход 4-20 мА)')
ws3.cell(row=row, column=1).font = sub_header_font
ws3.cell(row=row, column=1).fill = sub_header_fill
ws3.cell(row=row, column=1).alignment = center
for c in range(1, 5):
    ws3.cell(row=row, column=c).border = border_all

row = 4
for i, h in enumerate(['Клемма', 'Обозначение', 'От кабеля', 'Сигнал'], 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 1, 4)

dcs_data = [
    (1, 'T1 Poten+', 'Кабель 5, пара 1-1', 'Потенциал T1 (4-20 мА)'),
    (2, 'T1 Poten-', 'Кабель 5, пара 1-2', ''),
    (3, 'T1 pH+', 'Кабель 5, пара 2-1', 'pH / Концентрация T1 (4-20 мА)'),
    (4, 'T1 pH-', 'Кабель 5, пара 2-2', ''),
    (5, 'T2 Poten+', 'Кабель 6, пара 1-1', 'Потенциал T2 (4-20 мА)'),
    (6, 'T2 Poten-', 'Кабель 6, пара 1-2', ''),
    (7, 'T2 pH+', 'Кабель 6, пара 2-1', 'pH / Концентрация T2 (4-20 мА)'),
    (8, 'T2 pH-', 'Кабель 6, пара 2-2', ''),
]
for (term, label, cable, signal) in dcs_data:
    row += 1
    ws3.cell(row=row, column=1, value=term)
    ws3.cell(row=row, column=2, value=label)
    ws3.cell(row=row, column=3, value=cable)
    ws3.cell(row=row, column=4, value=signal)
    style_row(ws3, row, 1, 4)
    ws3.cell(row=row, column=1).alignment = center

# ── Connection Box ──
row += 2
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value='Connection Box (Sheet 5, XT)')
ws3.cell(row=row, column=1).font = sub_header_font
ws3.cell(row=row, column=1).fill = sub_header_fill
ws3.cell(row=row, column=1).alignment = center
for c in range(1, 5):
    ws3.cell(row=row, column=c).border = border_all

row += 1
for i, h in enumerate(['Клемма', 'Надпись', 'Вход (от)', 'Выход (к)'], 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 1, 4)

cb_data = [
    (1, '+24V', 'Кабель 9 (U+)', 'Кабель 3 → T1 (U1+)'),
    (2, '+24V', 'Кабель 9 (U+)', 'Кабель 4 → T2 (U2+)'),
    (3, '-24V', 'Кабель 9 (U-)', 'Кабель 3 → T1 (U1-)'),
    (4, '-24V', 'Кабель 9 (U-)', 'Кабель 4 → T2 (U2-)'),
    (5, 'Data+', 'Кабель 9 (D+)', 'Кабель 3 → T1 (Data1+)'),
    (6, 'Data+', 'Кабель 9 (D+)', 'Кабель 4 → T2 (Data2+)'),
    (7, 'Data-', 'Кабель 9 (D-)', 'Кабель 3 → T1 (Data1-)'),
    (8, 'Data-', 'Кабель 9 (D-)', 'Кабель 4 → T2 (Data2-)'),
]
for (term, label, inp, out) in cb_data:
    row += 1
    ws3.cell(row=row, column=1, value=term)
    ws3.cell(row=row, column=2, value=label)
    ws3.cell(row=row, column=3, value=inp)
    ws3.cell(row=row, column=4, value=out)
    style_row(ws3, row, 1, 4)
    ws3.cell(row=row, column=1).alignment = center

# ── Transmitter XT2 (typical) ──
row += 2
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1,
         value='Клеммник XT2 Трансмиттера (типовой, Sheets 2-3)')
ws3.cell(row=row, column=1).font = sub_header_font
ws3.cell(row=row, column=1).fill = sub_header_fill
ws3.cell(row=row, column=1).alignment = center
for c in range(1, 5):
    ws3.cell(row=row, column=c).border = border_all

row += 1
for i, h in enumerate(['Клемма', 'Надпись', 'Кабель', 'Сигнал'], 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 1, 4)

xt2_data = [
    (1, 'Supply+', 'Каб. 3/4 (от ConnBox)', '+24VDC'),
    (2, 'Supply-', 'Каб. 3/4 (от ConnBox)', '-24VDC (GND)'),
    (3, 'Data+', 'Каб. 3/4 (от ConnBox)', 'RS-485 A'),
    (4, 'Data-', 'Каб. 3/4 (от ConnBox)', 'RS-485 B'),
    (5, 'Potential+', 'Каб. 5/6 (к DCS)', '4-20 мА (QS1 OUT)'),
    (6, 'Potential-', 'Каб. 5/6 (к DCS)', '4-20 мА (QS1 GND)'),
    (7, 'Conc/pH+', 'Каб. 5/6 (к DCS)', '4-20 мА (QS2 OUT)'),
    (8, 'Conc/pH-', 'Каб. 5/6 (к DCS)', '4-20 мА (QS2 GND)'),
]
for (term, label, cable, signal) in xt2_data:
    row += 1
    ws3.cell(row=row, column=1, value=term)
    ws3.cell(row=row, column=2, value=label)
    ws3.cell(row=row, column=3, value=cable)
    ws3.cell(row=row, column=4, value=signal)
    style_row(ws3, row, 1, 4)
    ws3.cell(row=row, column=1).alignment = center


# ═══════════════════════════════════════════════════════════════════
# Sheet 4: СВОДКА МАТЕРИАЛОВ (Material Summary)
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Сводка материалов")
ws4.sheet_properties.tabColor = 'C00000'

col_w4 = [36, 14, 16, 12]
for i, w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells('A1:D1')
ws4['A1'] = 'СВОДКА МЕТРАЖА И МАТЕРИАЛОВ / MATERIAL SUMMARY'
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center')

row = 3
for i, h in enumerate(['Тип кабеля', 'Кабели №', 'Длина (шт.), м', 'Итого, м'], 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, 1, 4)

materials = [
    ('4×1.0 mm² с экраном', '1, 2', '15 + 15', 30),
    ('2 пары×1.5 mm² витая пара', '3, 4, 9', '15 + 15 + 250', 280),
    ('2 пары×1.5 mm² с экраном', '5, 6', '200 + 200', 400),
    ('1C×2.5 mm² жёлто-зелёный', '7, 8', '50 + 50', 100),
    ('3 жилы×2.5 mm²', '10', '100', 100),
    ('3×0.75 mm²', '11', '3', 3),
    ('HDMI', '12', '3', 3),
]
for idx, (mtype, nums, lengths, total) in enumerate(materials):
    row = 4 + idx
    ws4.cell(row=row, column=1, value=mtype)
    ws4.cell(row=row, column=2, value=nums)
    ws4.cell(row=row, column=3, value=lengths)
    ws4.cell(row=row, column=4, value=total)
    style_row(ws4, row, 1, 4)
    ws4.cell(row=row, column=1).alignment = left_wrap
    ws4.cell(row=row, column=4).alignment = center

# Total
row = 4 + len(materials)
ws4.merge_cells(f'A{row}:C{row}')
ws4.cell(row=row, column=1, value='ВСЕГО / TOTAL')
ws4.cell(row=row, column=4, value=f'=SUM(D4:D{row-1})')
style_row(ws4, row, 1, 4, font=sub_header_font, fill=sub_header_fill)
ws4.cell(row=row, column=1).alignment = Alignment(horizontal='right',
                                                    vertical='center')
ws4.cell(row=row, column=4).alignment = center


# ═══════════════════════════════════════════════════════════════════
# Sheet 5: ЗАМЕЧАНИЯ (Notes)
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Замечания")
ws5.sheet_properties.tabColor = 'FF0000'

col_w5 = [6, 10, 40, 40]
for i, w in enumerate(col_w5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells('A1:D1')
ws5['A1'] = 'ЗАМЕЧАНИЯ ПО ЭЛЕКТРИЧЕСКОЙ СХЕМЕ 421457.103E1'
ws5['A1'].font = Font(name='Arial', size=14, bold=True, color='CC0000')
ws5['A1'].alignment = Alignment(horizontal='center')

row = 3
for i, h in enumerate(['№', 'Лист', 'Ошибка', 'Исправление'], 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, 1, 4)

notes = [
    (1, 'Листы 2,3,4',
     'S/N и ADDRESS перепутаны местами на всех трёх листах трансмиттеров.\n'
     'Например, Лист 2: написано S/N 40, ADDRESS 32',
     'Поменять местами:\n'
     'Лист 2: S/N 32, ADDRESS 40\n'
     'Лист 3: S/N 34, ADDRESS 41\n'
     'Лист 4: S/N 36, ADDRESS 42'),
    (2, 'Лист 5',
     'Кабель от Connection box к Трансмиттеру №2 обозначен "6".\n'
     'Кабель 6 = 4-20мА от T2 к DCS (200м).',
     'Исправить номер кабеля на "4".\n'
     'Кабель 4 = 24VDC/RS485 от ConnBox к T2 (15м).'),
    (3, 'Лист 5',
     'Кабель питания 230VAC к SCADA box обозначен "8".\n'
     'Кабель 8 = заземление Трансмиттера №2 (50м).',
     'Исправить номер кабеля на "10".\n'
     'Кабель 10 = 230VAC питание (100м, 3ж×2.5мм²).'),
]
for (num, sheet, error, fix) in notes:
    row += 1
    ws5.cell(row=row, column=1, value=num)
    ws5.cell(row=row, column=2, value=sheet)
    ws5.cell(row=row, column=3, value=error)
    ws5.cell(row=row, column=4, value=fix)
    style_row(ws5, row, 1, 4, fill=error_fill)
    ws5.cell(row=row, column=1).alignment = center
    ws5.cell(row=row, column=2).alignment = center
    ws5.cell(row=row, column=3).alignment = left_top
    ws5.cell(row=row, column=4).alignment = left_top
    ws5.row_dimensions[row].height = 65

# ═══════════════════════════════════════════════════════════════════
# Sheet 6: РАСЧЁТ ПАДЕНИЯ НАПРЯЖЕНИЯ (Voltage Drop Calculation)
# ═══════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Падение напряжения")
ws6.sheet_properties.tabColor = 'FF6600'

warn_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
crit_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
info_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')

col_w6 = [40, 14, 14, 14, 14, 14]
for i, w in enumerate(col_w6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.merge_cells('A1:F1')
ws6['A1'] = 'РАСЧЁТ ПАДЕНИЯ НАПРЯЖЕНИЯ В КАБЕЛЯХ ПИТАНИЯ / VOLTAGE DROP CALCULATION'
ws6['A1'].font = title_font
ws6['A1'].alignment = Alignment(horizontal='center')

ws6.merge_cells('A2:F2')
ws6['A2'] = 'DEEPAK K1 — GP Plant  |  421457.103E1'
ws6['A2'].font = Font(name='Arial', size=10, italic=True)
ws6['A2'].alignment = Alignment(horizontal='center')

# ── Section: Схема питания ──
row = 4
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1, value='СХЕМА ЦЕПИ ПИТАНИЯ ТРАНСМИТТЕРОВ')
ws6.cell(row=row, column=1).font = sub_header_font
ws6.cell(row=row, column=1).fill = sub_header_fill
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all

row = 5
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1,
    value='SCADA Ekor Box (24VDC) ──[ Cable 9: 250м, 1.5мм² ]──> Connection Box '
          '──[ Cable 3: 15м ]──> Трансмиттер №1')
ws6.cell(row=row, column=1).font = Font(name='Consolas', size=10)
ws6.cell(row=row, column=1).alignment = left_wrap
ws6.row_dimensions[row].height = 18

row = 6
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1,
    value='                                                       '
          '                   └──[ Cable 4: 15м ]──> Трансмиттер №2')
ws6.cell(row=row, column=1).font = Font(name='Consolas', size=10)
ws6.cell(row=row, column=1).alignment = left_wrap

row = 7
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1,
    value='ВАЖНО: Cable 9 несёт суммарный ток ОБОИХ трансмиттеров!')
ws6.cell(row=row, column=1).font = note_font
ws6.cell(row=row, column=1).fill = error_fill
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all

# ── Section: Исходные данные ──
row = 9
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1, value='ИСХОДНЫЕ ДАННЫЕ')
ws6.cell(row=row, column=1).font = sub_header_font
ws6.cell(row=row, column=1).fill = sub_header_fill
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all

input_data = [
    ('Напряжение источника (SCADA Ekor Box)', '24.0 В'),
    ('Удельное сопротивление меди при 20°C (ρ₂₀)', '0.0175 Ом·мм²/м'),
    ('Температурный коэффициент меди (α)', '0.00393 1/°C'),
    ('Cable 9: длина / сечение', '250 м / 1.5 мм²'),
    ('Cable 3: длина / сечение', '15 м / 1.5 мм²'),
    ('Cable 4: длина / сечение', '15 м / 1.5 мм²'),
    ('Допустимый диапазон питания трансмиттера', '18...30 В'),
    ('Рекомендуемый минимум (с запасом на пульсации)', '20 В'),
]
for label, val in input_data:
    row += 1
    ws6.cell(row=row, column=1, value=label)
    ws6.cell(row=row, column=1).font = normal_font
    ws6.cell(row=row, column=1).alignment = left_wrap
    ws6.merge_cells(f'B{row}:F{row}')
    ws6.cell(row=row, column=2, value=val)
    ws6.cell(row=row, column=2).font = sub_header_font
    ws6.cell(row=row, column=2).alignment = center
    for c in range(1, 7):
        ws6.cell(row=row, column=c).border = border_all

# ── Section: Потребление тока ──
row += 2
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1,
    value='ПОТРЕБЛЕНИЕ ТОКА ТРАНСМИТТЕРА (из документации K1)')
ws6.cell(row=row, column=1).font = sub_header_font
ws6.cell(row=row, column=1).fill = sub_header_fill
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all

current_breakdown = [
    ('Паспортное: "Power consumption of the transmitter: 8 W (0.35A)"',
     '350 мА (8 Вт / 24 В)'),
    ('I = P / U = 8 Вт / 24 В = 0.333 А (округлено до 0.35 А)', ''),
    ('Суммарный ток в Cable 9 (два трансмиттера)', '700 мА'),
]
for label, val in current_breakdown:
    row += 1
    ws6.cell(row=row, column=1, value=label)
    ws6.cell(row=row, column=1).font = small_font
    ws6.cell(row=row, column=1).alignment = left_wrap
    ws6.merge_cells(f'B{row}:F{row}')
    ws6.cell(row=row, column=2, value=val)
    ws6.cell(row=row, column=2).font = normal_font
    ws6.cell(row=row, column=2).alignment = center
    for c in range(1, 7):
        ws6.cell(row=row, column=c).border = border_all

# ── Section: Сопротивление кабелей ──
row += 2
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1, value='СОПРОТИВЛЕНИЕ КАБЕЛЕЙ (одна жила, Ом)')
ws6.cell(row=row, column=1).font = sub_header_font
ws6.cell(row=row, column=1).fill = sub_header_fill
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all

row += 1
resist_headers = ['Кабель', 'R@20°C', 'R@35°C', 'R@50°C', 'R@65°C', 'R петли@50°C']
for i, h in enumerate(resist_headers, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, 1, 6)

rho_20 = 0.0175
alpha_cu = 0.00393
cable_params = [
    ('Cable 9 (250м, 1.5мм²)', 250, 1.5),
    ('Cable 3 (15м, 1.5мм²)', 15, 1.5),
    ('Cable 4 (15м, 1.5мм²)', 15, 1.5),
]
for (name, length, section) in cable_params:
    row += 1
    ws6.cell(row=row, column=1, value=name)
    for idx, temp in enumerate([20, 35, 50, 65]):
        rho_t = rho_20 * (1 + alpha_cu * (temp - 20))
        r = rho_t * length / section
        ws6.cell(row=row, column=2 + idx, value=round(r, 3))
    # Loop resistance at 50°C
    rho_50 = rho_20 * (1 + alpha_cu * 30)
    r_loop = 2 * rho_50 * length / section
    ws6.cell(row=row, column=6, value=round(r_loop, 3))
    style_row(ws6, row, 1, 6)
    ws6.cell(row=row, column=1).alignment = left_wrap

# ═══════════════════════════════════════════════════════════════════
# MAIN CALCULATION TABLE
# ═══════════════════════════════════════════════════════════════════
row += 2
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1,
    value='СВОДНАЯ ТАБЛИЦА: НАПРЯЖЕНИЕ НА КЛЕММАХ ТРАНСМИТТЕРА (В)')
ws6.cell(row=row, column=1).font = title_font
ws6.cell(row=row, column=1).fill = PatternFill(start_color='FFC000',
    end_color='FFC000', fill_type='solid')
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all
ws6.row_dimensions[row].height = 30

row += 1
summary_headers = ['Режим работы', 'I транс., мА', '20°C', '35°C', '50°C', '65°C']
for i, h in enumerate(summary_headers, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, 1, 6)

scenarios = [
    ('Минимальный (пуск, прогрев)', 200),
    ('Номинальный (выходы ~12мА)', 300),
    ('Паспортный (8 Вт / 24 В = 0.35 А)', 350),
    ('Максимальный (с запасом 10%)', 385),
]

for (scenario_name, I_mA) in scenarios:
    row += 1
    I_tx = I_mA / 1000.0
    I_total_cab9 = 2 * I_tx
    ws6.cell(row=row, column=1, value=scenario_name)
    ws6.cell(row=row, column=1).font = normal_font
    ws6.cell(row=row, column=1).alignment = left_wrap
    ws6.cell(row=row, column=2, value=I_mA)
    ws6.cell(row=row, column=2).font = sub_header_font
    ws6.cell(row=row, column=2).alignment = center

    for idx, temp in enumerate([20, 35, 50, 65]):
        rho_t = rho_20 * (1 + alpha_cu * (temp - 20))
        R9 = rho_t * 250 / 1.5
        R3 = rho_t * 15 / 1.5
        dU = 2 * R9 * I_total_cab9 + 2 * R3 * I_tx
        V = 24.0 - dU
        cell = ws6.cell(row=row, column=3 + idx, value=round(V, 2))
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.alignment = center
        if V < 18.0:
            cell.fill = crit_fill
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        elif V < 20.0:
            cell.fill = warn_fill
        else:
            cell.fill = ok_fill

    for c in range(1, 7):
        ws6.cell(row=row, column=c).border = border_all
    ws6.row_dimensions[row].height = 24

# ── Legend ──
row += 1
ws6.cell(row=row, column=1, value='Цветовая шкала:')
ws6.cell(row=row, column=1).font = small_font
ws6.cell(row=row, column=3, value='> 20 В')
ws6.cell(row=row, column=3).fill = ok_fill
ws6.cell(row=row, column=3).alignment = center
ws6.cell(row=row, column=3).border = border_all
ws6.cell(row=row, column=4, value='< 20 В')
ws6.cell(row=row, column=4).fill = warn_fill
ws6.cell(row=row, column=4).alignment = center
ws6.cell(row=row, column=4).border = border_all
ws6.cell(row=row, column=5, value='< 18 В')
ws6.cell(row=row, column=5).fill = crit_fill
ws6.cell(row=row, column=5).font = Font(name='Arial', size=10, color='FFFFFF')
ws6.cell(row=row, column=5).alignment = center
ws6.cell(row=row, column=5).border = border_all

# ═══════════════════════════════════════════════════════════════════
# DETAILED BREAKDOWN (worst case)
# ═══════════════════════════════════════════════════════════════════
row += 2
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1,
    value='ДЕТАЛЬНЫЙ РАСЧЁТ — ПАСПОРТНЫЙ РЕЖИМ (350 мА = 8 Вт, 65°C)')
ws6.cell(row=row, column=1).font = sub_header_font
ws6.cell(row=row, column=1).fill = PatternFill(start_color='FFC000',
    end_color='FFC000', fill_type='solid')
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all

# Calculate worst case (passport: 8W / 24V = 0.35A)
I_tx_w = 0.350
I_cab9_w = 2 * I_tx_w
rho_65 = rho_20 * (1 + alpha_cu * 45)
R9_w = rho_65 * 250 / 1.5
R3_w = rho_65 * 15 / 1.5
dU_cab9_w = 2 * R9_w * I_cab9_w
dU_cab3_w = 2 * R3_w * I_tx_w
V_connbox_w = 24.0 - dU_cab9_w
V_T1_w = 24.0 - dU_cab9_w - dU_cab3_w
P_cab9_w = dU_cab9_w * I_cab9_w
P_cab3_w = dU_cab3_w * I_tx_w

detail_rows = [
    ('Ток одного трансмиттера', f'{I_tx_w*1000:.0f} мА'),
    ('Суммарный ток в Cable 9 (оба трансмиттера)', f'{I_cab9_w*1000:.0f} мА'),
    ('R(Cable 9) при 65°C — одна жила', f'{R9_w:.3f} Ом'),
    ('R(Cable 9) при 65°C — петля (туда+обратно)', f'{2*R9_w:.3f} Ом'),
    ('R(Cable 3/4) при 65°C — петля', f'{2*R3_w:.3f} Ом'),
    ('', ''),
    ('ΔU на Cable 9 (250 м)', f'{dU_cab9_w:.2f} В'),
    ('ΔU на Cable 3/4 (15 м)', f'{dU_cab3_w:.2f} В'),
    ('ИТОГО падение напряжения', f'{dU_cab9_w + dU_cab3_w:.2f} В  ({(dU_cab9_w + dU_cab3_w)/24*100:.1f}%)'),
    ('', ''),
    ('V на Connection Box', f'{V_connbox_w:.2f} В'),
    ('V НА ТРАНСМИТТЕРЕ', f'{V_T1_w:.2f} В'),
    ('', ''),
    ('Мощность потерь в Cable 9', f'{P_cab9_w:.2f} Вт'),
    ('Мощность потерь в Cable 3/4', f'{P_cab3_w:.3f} Вт'),
    ('Суммарные тепловые потери', f'{P_cab9_w + P_cab3_w:.2f} Вт'),
]

for (label, val) in detail_rows:
    row += 1
    ws6.cell(row=row, column=1, value=label)
    ws6.merge_cells(f'B{row}:F{row}')
    ws6.cell(row=row, column=2, value=val)
    font = sub_header_font if 'ИТОГО' in label or 'НА ТРАНСМИТТЕРЕ' in label else normal_font
    fill_d = error_fill if 'ИТОГО' in label or 'НА ТРАНСМИТТЕРЕ' in label else white_fill
    ws6.cell(row=row, column=1).font = font
    ws6.cell(row=row, column=1).alignment = left_wrap
    ws6.cell(row=row, column=2).font = font
    ws6.cell(row=row, column=2).alignment = center
    if label:
        for c in range(1, 7):
            ws6.cell(row=row, column=c).border = border_all
        ws6.cell(row=row, column=1).fill = fill_d
        ws6.cell(row=row, column=2).fill = fill_d

# ═══════════════════════════════════════════════════════════════════
# RECOMMENDATIONS
# ═══════════════════════════════════════════════════════════════════
row += 2
ws6.merge_cells(f'A{row}:F{row}')
ws6.cell(row=row, column=1, value='РЕКОМЕНДАЦИИ / RECOMMENDATIONS')
ws6.cell(row=row, column=1).font = title_font
ws6.cell(row=row, column=1).fill = sub_header_fill
ws6.cell(row=row, column=1).alignment = center
for c in range(1, 7):
    ws6.cell(row=row, column=c).border = border_all

recommendations = [
    ('1. Увеличить сечение Cable 9 до 2.5 мм²',
     f'V на трансмиттере (65°C, 350мА)',
     round(24.0 - 2 * (rho_65 * 250 / 2.5) * I_cab9_w - dU_cab3_w, 2)),
    ('2. Повысить напряжение источника до 27 В',
     f'V на трансмиттере (65°C, 350мА)',
     round(V_T1_w + 3, 2)),
    ('3. Сократить Cable 9 до 100 м (перенести SCADA box)',
     f'V на трансмиттере (65°C, 350мА)',
     round(24.0 - 2 * (rho_65 * 100 / 1.5) * I_cab9_w - dU_cab3_w, 2)),
    ('4. Отдельный БП 24VDC у Connection Box',
     f'V на трансмиттере (без Cable 9)',
     round(24.0 - dU_cab3_w, 2)),
]

row += 1
for i, h in enumerate(['Рекомендация', 'Параметр', 'V, В',
                        'Падение, В', 'Статус'], 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, 1, 5)

for (rec, param, v_result) in recommendations:
    row += 1
    ws6.cell(row=row, column=1, value=rec)
    ws6.cell(row=row, column=1).font = normal_font
    ws6.cell(row=row, column=1).alignment = left_wrap
    ws6.cell(row=row, column=2, value=param)
    ws6.cell(row=row, column=2).font = small_font
    ws6.cell(row=row, column=2).alignment = left_wrap
    ws6.cell(row=row, column=3, value=v_result)
    ws6.cell(row=row, column=3).font = Font(name='Arial', size=11, bold=True)
    ws6.cell(row=row, column=3).alignment = center
    ws6.cell(row=row, column=3).fill = ok_fill
    ws6.cell(row=row, column=4, value=round(24.0 - v_result, 2))
    ws6.cell(row=row, column=4).font = normal_font
    ws6.cell(row=row, column=4).alignment = center
    ws6.cell(row=row, column=5, value='OK')
    ws6.cell(row=row, column=5).font = Font(name='Arial', size=11, bold=True,
                                              color='006100')
    ws6.cell(row=row, column=5).fill = ok_fill
    ws6.cell(row=row, column=5).alignment = center
    for c in range(1, 6):
        ws6.cell(row=row, column=c).border = border_all
    ws6.row_dimensions[row].height = 30


# ── Print setup ────────────────────────────────────────────────────
for ws in [ws1, ws2, ws3, ws4, ws5, ws6]:
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

# ── Save ───────────────────────────────────────────────────────────
out_path = '/home/user/46clod/K1_Cable_Journal.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")
print("Sheets: Кабельный журнал, Спецификация жил, Клеммные карты, "
      "Сводка материалов, Замечания")
